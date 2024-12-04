from django.shortcuts import render, redirect
from django.contrib import messages
from .models import LeaveRequest, Employee
from django.contrib.auth.decorators import login_required

@login_required
def create_leave_request(request):
    if request.method == 'POST':
        # Get the data from the form
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        number_of_days = int(request.POST.get('number_of_days'))
        contact_address = request.POST.get('contact_address_during_leave', '')  # Default to empty string if not provided
        leave_grant_requested = request.POST.get('leave_grant_requested', '')
        
        # Get the employee object associated with the logged-in user
        employee = request.user.employee  # Assuming the logged-in user has an associated Employee object

        # Create the leave request object and associate the user
        leave_request = LeaveRequest(
            employee=employee,
            start_date=start_date,
            end_date=end_date,
            number_of_days=number_of_days,
            contact_address_during_leave=contact_address,
            leave_grant_requested=leave_grant_requested,
            user=request.user  # Set the logged-in user to the 'user' field
        )
        leave_request.save()

        # Show a success message
        messages.success(request, 'Your leave request has been submitted and is pending approval.')
        return redirect('homepage')
    
    # Fetch all leave requests to display
    leave_requests = LeaveRequest.objects.filter(user=request.user)  # Only show the logged-in user's leave requests

    leave_requests_data = []
    for leave in leave_requests:
        # Check if the approval exists for 'Head of Department'
        head_approved = leave.leaveapproval_set.filter(
            approver_role__role_name='Head of Department',  # Match database role name
            approval_status='Approved'  # Use the correct field name
        ).exists()

        leave_requests_data.append({
            'id': leave.id,
            'employee_name': leave.employee.name,
            'start_date': leave.start_date,
            'end_date': leave.end_date,
            'status': "Approved by Head" if head_approved else "Pending",
            'download_enabled': head_approved,  # Enable only if Head approved
        })

    return render(request, 'request_leave.html', {'leave_requests': leave_requests_data})

def create_leave_request_sp(request):
    if request.method == 'POST':
        # Get the data from the form
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        number_of_days = int(request.POST.get('number_of_days'))
        contact_address = request.POST.get('contact_address_during_leave', '')  # Default to empty string if not provided
        leave_grant_requested = request.POST.get('leave_grant_requested', '')
        
        # Get the employee object associated with the logged-in user
        employee = request.user.employee  # Assuming the logged-in user has an associated Employee object

        # Create the leave request object and associate the user
        leave_request = LeaveRequest(
            employee=employee,
            start_date=start_date,
            end_date=end_date,
            number_of_days=number_of_days,
            contact_address_during_leave=contact_address,
            leave_grant_requested=leave_grant_requested,
            user=request.user  # Set the logged-in user to the 'user' field
        )
        leave_request.save()

        # Show a success message
        messages.success(request, 'Your leave request has been submitted and is pending approval.')
        return redirect('homepage')
    
    # Fetch all leave requests to display
    leave_requests = LeaveRequest.objects.filter(user=request.user)  # Only show the logged-in user's leave requests

    leave_requests_data = []
    for leave in leave_requests:
        # Check if the approval exists for 'Head of Department'
        head_approved = leave.leaveapproval_set.filter(
            approver_role__role_name='Head of Department',  # Match database role name
            approval_status='Approved'  # Use the correct field name
        ).exists()

        leave_requests_data.append({
            'id': leave.id,
            'employee_name': leave.employee.name,
            'start_date': leave.start_date,
            'end_date': leave.end_date,
            'status': "Approved by Head" if head_approved else "Pending",
            'download_enabled': head_approved,  # Enable only if Head approved
        })

    return render(request, 'sp/request_leave.html', {'leave_requests': leave_requests_data})

def create_leave_request_hr(request):
    if request.method == 'POST':
        # Get the data from the form
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        number_of_days = int(request.POST.get('number_of_days'))
        contact_address = request.POST.get('contact_address_during_leave', '')  # Default to empty string if not provided
        leave_grant_requested = request.POST.get('leave_grant_requested', '')
        
        # Get the employee object associated with the logged-in user
        employee = request.user.employee  # Assuming the logged-in user has an associated Employee object

        # Create the leave request object and associate the user
        leave_request = LeaveRequest(
            employee=employee,
            start_date=start_date,
            end_date=end_date,
            number_of_days=number_of_days,
            contact_address_during_leave=contact_address,
            leave_grant_requested=leave_grant_requested,
            user=request.user  # Set the logged-in user to the 'user' field
        )
        leave_request.save()

        # Show a success message
        messages.success(request, 'Your leave request has been submitted and is pending approval.')
        return redirect('homepage')
    
    # Fetch all leave requests to display
    leave_requests = LeaveRequest.objects.filter(user=request.user)  # Only show the logged-in user's leave requests

    leave_requests_data = []
    for leave in leave_requests:
        # Check if the approval exists for 'Head of Department'
        head_approved = leave.leaveapproval_set.filter(
            approver_role__role_name='Head of Department',  # Match database role name
            approval_status='Approved'  # Use the correct field name
        ).exists()

        leave_requests_data.append({
            'id': leave.id,
            'employee_name': leave.employee.name,
            'start_date': leave.start_date,
            'end_date': leave.end_date,
            'status': "Approved by Head" if head_approved else "Pending",
            'download_enabled': head_approved,  # Enable only if Head approved
        })

    return render(request, 'hr/request_leave.html', {'leave_requests': leave_requests_data})

def create_leave_request_hd(request):
    if request.method == 'POST':
        # Get the data from the form
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        number_of_days = int(request.POST.get('number_of_days'))
        contact_address = request.POST.get('contact_address_during_leave', '')  # Default to empty string if not provided
        leave_grant_requested = request.POST.get('leave_grant_requested', '')
        
        # Get the employee object associated with the logged-in user
        employee = request.user.employee  # Assuming the logged-in user has an associated Employee object

        # Create the leave request object and associate the user
        leave_request = LeaveRequest(
            employee=employee,
            start_date=start_date,
            end_date=end_date,
            number_of_days=number_of_days,
            contact_address_during_leave=contact_address,
            leave_grant_requested=leave_grant_requested,
            user=request.user  # Set the logged-in user to the 'user' field
        )
        leave_request.save()

        # Show a success message
        messages.success(request, 'Your leave request has been submitted and is pending approval.')
        return redirect('homepage')
    
    # Fetch all leave requests to display
    leave_requests = LeaveRequest.objects.filter(user=request.user)  # Only show the logged-in user's leave requests

    leave_requests_data = []
    for leave in leave_requests:
        # Check if the approval exists for 'Head of Department'
        head_approved = leave.leaveapproval_set.filter(
            approver_role__role_name='Head of Department',  # Match database role name
            approval_status='Approved'  # Use the correct field name
        ).exists()

        leave_requests_data.append({
            'id': leave.id,
            'employee_name': leave.employee.name,
            'start_date': leave.start_date,
            'end_date': leave.end_date,
            'status': "Approved by Head" if head_approved else "Pending",
            'download_enabled': head_approved,  # Enable only if Head approved
        })

    return render(request, 'hd/request_leave.html', {'leave_requests': leave_requests_data})

def create_leave_request_adm(request):
    if request.method == 'POST':
        # Get the data from the form
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        number_of_days = int(request.POST.get('number_of_days'))
        contact_address = request.POST.get('contact_address_during_leave', '')  # Default to empty string if not provided
        leave_grant_requested = request.POST.get('leave_grant_requested', '')
        
        # Get the employee object associated with the logged-in user
        employee = request.user.employee  # Assuming the logged-in user has an associated Employee object

        # Create the leave request object and associate the user
        leave_request = LeaveRequest(
            employee=employee,
            start_date=start_date,
            end_date=end_date,
            number_of_days=number_of_days,
            contact_address_during_leave=contact_address,
            leave_grant_requested=leave_grant_requested,
            user=request.user  # Set the logged-in user to the 'user' field
        )
        leave_request.save()

        # Show a success message
        messages.success(request, 'Your leave request has been submitted and is pending approval.')
        return redirect('homepage')
    
    # Fetch all leave requests to display
    leave_requests = LeaveRequest.objects.filter(user=request.user)  # Only show the logged-in user's leave requests

    leave_requests_data = []
    for leave in leave_requests:
        # Check if the approval exists for 'Head of Department'
        head_approved = leave.leaveapproval_set.filter(
            approver_role__role_name='Head of Department',  # Match database role name
            approval_status='Approved'  # Use the correct field name
        ).exists()

        leave_requests_data.append({
            'id': leave.id,
            'employee_name': leave.employee.name,
            'start_date': leave.start_date,
            'end_date': leave.end_date,
            'status': "Approved by Head" if head_approved else "Pending",
            'download_enabled': head_approved,  # Enable only if Head approved
        })

    return render(request, 'admin_z/request_leave.html', {'leave_requests': leave_requests_data})



# @login_required
# def hr_approval(request, leave_request_id):
#     leave_request = get_object_or_404(LeaveRequest, id=leave_request_id)

#     # Check if the user has the HR role
#     if request.user.employee.post and request.user.employee.post.role_name in ['HR']:  # Only HR can approve
#         if request.method == 'POST':
#             approval_status = request.POST.get('approval_status')
#             comments = request.POST.get('comments', '')

#             if approval_status == 'Rejected':
#                 leave_request.status = 'Rejected'
#                 leave_request.save()
#                 messages.error(request, 'Leave request has been rejected.')
#             else:
#                 leave_request.status = 'HR Approved'
#                 leave_request.save()

#                 # Check if an HR approval exists, update it or create a new one
#                 hr_approval = leave_request.leaveapproval_set.filter(approver_role__role_name='HR').first()
                
#                 if hr_approval:
#                     # If HR approval exists, update it
#                     hr_approval.approval_status = 'Approved'
#                     hr_approval.comments = comments
#                     hr_approval.save()
#                     messages.success(request, 'Leave request updated and forwarded to Supervisor.')
#                 else:
#                     LeaveApproval.objects.create(
#                         leave_request=leave_request,
#                         approver=request.user.employee,  # Assuming 'employee' is related to 'User'
#                         approver_role=request.user.employee.post,
#                         approval_status='Approved',
#                         comments=comments
#                     )
#                     messages.success(request, 'Leave request approved and forwarded to Supervisor.')

#             return redirect('hr_dashboard')

#         return render(request, 'hr/hr_approval.html', {'leave_request': leave_request})

#     else:
#         messages.error(request, 'You are not authorized to approve leave requests.')
#         return redirect('hr_dashboard')
# @login_required
# def supervisor_approval(request, leave_request_id):
#     leave_request = get_object_or_404(LeaveRequest, id=leave_request_id)

#     # Supervisor can approve only if HR has approved
#     if leave_request.status == 'HR Approved' and request.user.employee.post and request.user.employee.post.role_name == 'Supervisor':
#         # Get the HR and Supervisor approvals
#         hr_approval = leave_request.leaveapproval_set.filter(approver_role__role_name='HR').first()
#         supervisor_approval = leave_request.leaveapproval_set.filter(approver_role__role_name='Supervisor').first()

#         if request.method == 'POST':
#             approval_status = request.POST.get('approval_status')
#             comments = request.POST.get('comments', '')

#             if approval_status == 'Rejected':
#                 leave_request.status = 'Rejected'
#                 leave_request.save()
#                 messages.error(request, 'Leave request has been rejected by Supervisor.')
#             else:
#                 leave_request.status = 'Supervisor Approved'
#                 leave_request.save()

#                 LeaveApproval.objects.create(
#                     leave_request=leave_request,
#                     approver=request.user.employee,  # Assuming 'employee' is related to 'User'
#                     approver_role=request.user.employee.post,
#                     approval_status='Approved',
#                     comments=comments
#                 )
#                 messages.success(request, 'Leave request approved and forwarded to Head of Section.')

#             return redirect('supervisor_dashboard')

#         return render(request, 'sp/supervisor_approval.html', {
#             'leave_request': leave_request,
#             'hr_approval': hr_approval,
#             'supervisor_approval': supervisor_approval
#         })

#     else:
#         messages.error(request, 'You are not authorized to approve this leave request or it is not at the correct stage.')
#         return redirect('supervisor_dashboard')
# @login_required
# def head_of_section_approval(request, leave_request_id):
#     leave_request = get_object_or_404(LeaveRequest, id=leave_request_id)

#     # Head of Section can approve the leave request only if Supervisor has approved it
#     if leave_request.status == 'Supervisor Approved' and request.user.employee.post and request.user.employee.post.role_name == 'Head of Department':
#         # Get the HR, Supervisor, and Head of Section approvals
#         hr_approval = leave_request.leaveapproval_set.filter(approver_role__role_name='HR').first()
#         supervisor_approval = leave_request.leaveapproval_set.filter(approver_role__role_name='Supervisor').first()
#         head_of_section_approval = leave_request.leaveapproval_set.filter(approver_role__role_name='Head of Department').first()

#         if request.method == 'POST':
#             approval_status = request.POST.get('approval_status')
#             comments = request.POST.get('comments', '')

#             if approval_status == 'Rejected':
#                 leave_request.status = 'Rejected'
#                 leave_request.save()
#                 messages.error(request, 'Leave request has been rejected by Head of Section.')
#             else:
#                 leave_request.status = 'Head Approved'
#                 leave_request.save()
#                 LeaveApproval.objects.create(
#                     leave_request=leave_request,
#                     approver=request.user.employee,
#                     approver_role=request.user.employee.post,
#                     approval_status='Approved',
#                     comments=comments
#                 )
#                 messages.success(request, 'Leave request approved successfully.')

#             return redirect('head_of_section_dashboard')

#         return render(request, 'hd/head_of_section_approval.html', {
#             'leave_request': leave_request,
#             'hr_approval': hr_approval,
#             'supervisor_approval': supervisor_approval,
#             'head_of_section_approval': head_of_section_approval,
#         })

#     else:
#         messages.error(request, 'Leave request cannot be approved by you at this stage.')
#         return redirect('head_of_section_dashboard')
from django.utils.timezone import now


@login_required
def hr_approval(request, leave_request_id):
    leave_request = get_object_or_404(LeaveRequest, id=leave_request_id)

    if request.user.employee.post and request.user.employee.post.role_name == 'HR':
        if request.method == 'POST':
            approval_status = request.POST.get('approval_status')
            comments = request.POST.get('comments', '')
            signature = request.FILES.get('signature')  # Handle signature upload

            if approval_status == 'Rejected':
                leave_request.status = 'Rejected'
                leave_request.save()
                messages.error(request, 'Leave request has been rejected.')
            else:
                leave_request.status = 'HR Approved'
                leave_request.save()

                # Update or create the HR approval entry
                hr_approval, created = LeaveApproval.objects.update_or_create(
                    leave_request=leave_request,
                    approver_role=request.user.employee.post,
                    defaults={
                        'approver': request.user.employee,
                        'approval_status': 'Approved',
                        'comments': comments,
                        'signature': signature,
                        'action_date': now()
                    }
                )
                messages.success(request, 'Leave request approved and forwarded to Supervisor.')

            return redirect('hr_dashboard')

        return render(request, 'hr/hr_approval.html', {'leave_request': leave_request})

    messages.error(request, 'You are not authorized to approve leave requests.')
    return redirect('hr_dashboard')

@login_required
def supervisor_approval(request, leave_request_id):
    leave_request = get_object_or_404(LeaveRequest, id=leave_request_id)

    if leave_request.status == 'HR Approved' and request.user.employee.post.role_name == 'Supervisor':
        if request.method == 'POST':
            approval_status = request.POST.get('approval_status')
            comments = request.POST.get('comments', '')
            signature = request.FILES.get('signature')  # Handle signature upload

            if approval_status == 'Rejected':
                leave_request.status = 'Rejected'
                leave_request.save()
                messages.error(request, 'Leave request has been rejected by Supervisor.')
            else:
                leave_request.status = 'Supervisor Approved'
                leave_request.save()

                LeaveApproval.objects.update_or_create(
                    leave_request=leave_request,
                    approver_role=request.user.employee.post,
                    defaults={
                        'approver': request.user.employee,
                        'approval_status': 'Approved',
                        'comments': comments,
                        'signature': signature,
                        'action_date': now()
                    }
                )
                messages.success(request, 'Leave request approved and forwarded to Head of Section.')

            return redirect('supervisor_dashboard')

        return render(request, 'sp/supervisor_approval.html', {'leave_request': leave_request})

    messages.error(request, 'You are not authorized to approve this leave request or it is not at the correct stage.')
    return redirect('supervisor_dashboard')

@login_required
def head_of_section_approval(request, leave_request_id):
    leave_request = get_object_or_404(LeaveRequest, id=leave_request_id)

    # Head of Section can approve only if Supervisor has approved
    if leave_request.status == 'Supervisor Approved' and request.user.employee.post and request.user.employee.post.role_name == 'Head of Department':
        if request.method == 'POST':
            approval_status = request.POST.get('approval_status')
            comments = request.POST.get('comments', '')
            signature = request.FILES.get('signature')  # Handle signature upload

            if approval_status == 'Rejected':
                leave_request.status = 'Rejected'
                leave_request.save()
                messages.error(request, 'Leave request has been rejected by Head of Section.')
            else:
                leave_request.status = 'Head Approved'
                leave_request.save()

                # Update or create Head of Section approval entry
                LeaveApproval.objects.update_or_create(
                    leave_request=leave_request,
                    approver_role=request.user.employee.post,
                    defaults={
                        'approver': request.user.employee,
                        'approval_status': 'Approved',
                        'comments': comments,
                        'signature': signature,
                        'action_date': now()
                    }
                )
                messages.success(request, 'Leave request has been approved successfully.')

            return redirect('head_of_section_dashboard')

        # Retrieve approval details for display (optional)
        hr_approval = leave_request.leaveapproval_set.filter(approver_role__role_name='HR').first()
        supervisor_approval = leave_request.leaveapproval_set.filter(approver_role__role_name='Supervisor').first()
        head_of_section_approval = leave_request.leaveapproval_set.filter(approver_role__role_name='Head of Department').first()

        return render(request, 'hd/head_of_section_approval.html', {
            'leave_request': leave_request,
            'hr_approval': hr_approval,
            'supervisor_approval': supervisor_approval,
            'head_of_section_approval': head_of_section_approval,
        })

    messages.error(request, 'Leave request cannot be approved by you at this stage.')
    return redirect('head_of_section_dashboard')


from django.shortcuts import render
from .models import LeaveRequest

@login_required
def leave_request_list(request):
    employee = request.user.employee
    leave_requests = LeaveRequest.objects.filter(employee=employee)

    return render(request, 'leave_request_list.html', {'leave_requests': leave_requests})

from django.core.paginator import Paginator
from django.shortcuts import render
from .models import LeaveRequest

def home(request):
    # Filter leave requests by the currently logged-in user
    leave_requests = LeaveRequest.objects.filter(user=request.user)

    leave_requests_data = []
    for leave in leave_requests:
        # Check if the approval exists for 'Head of Department'
        head_approved = leave.leaveapproval_set.filter(
            approver_role__role_name='Head of Department',  # Match database role name
            approval_status='Approved'  # Use the correct field name
        ).exists()

        leave_requests_data.append({
            'id': leave.id,
            'employee_name': leave.employee.name,
            'start_date': leave.start_date,
            'end_date': leave.end_date,
            'status': "Approved by Head" if head_approved else "Pending",
            'download_enabled': head_approved,  # Enable only if Head approved
        })
    
    # Paginate the leave requests data
    paginator = Paginator(leave_requests_data, 10)  # Show 10 leave requests per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'home.html', {'page_obj': page_obj})



from django.shortcuts import render
from django.urls import reverse



@login_required
def leave_requests_table(request):
    leave_requests = LeaveRequest.objects.all()

    leave_requests_data = []
    for leave in leave_requests:
        # Check if the approval exists for 'Head of Department'
        head_approved = leave.leaveapproval_set.filter(
            approver_role__role_name='Head of Department',  # Match database role name
            approval_status='Approved'  # Use the correct field name
        ).exists()

        leave_requests_data.append({
            'id': leave.id,
            'employee_name': leave.employee.name,
            'start_date': leave.start_date,
            'end_date': leave.end_date,
            'status': "Approved by Head" if head_approved else "Pending",
            'download_enabled': head_approved,  # Enable only if Head approved
        })

    return render(request, 'leave_requests_table.html', {'leave_requests': leave_requests_data})

from django.core.paginator import Paginator
from django.shortcuts import render

def leave_requests_table_hr(request):
    leave_requests = LeaveRequest.objects.all()

    leave_requests_data = []
    for leave in leave_requests:
        head_approved = leave.leaveapproval_set.filter(
            approver_role__role_name='Head of Department',
            approval_status='Approved'
        ).exists()

        leave_requests_data.append({
            'id': leave.id,
            'employee_name': leave.employee.name,
            'start_date': leave.start_date,
            'end_date': leave.end_date,
            'status': "Approved by Head" if head_approved else "Pending",
            'download_enabled': head_approved,
        })

    # Paginate the leave requests data
    paginator = Paginator(leave_requests_data, 10)  # Show 10 items per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'hr/leave_requests_table.html', {'page_obj': page_obj})


def leave_requests_table_sp(request):
    leave_requests = LeaveRequest.objects.all()

    leave_requests_data = []
    for leave in leave_requests:
        head_approved = leave.leaveapproval_set.filter(
            approver_role__role_name='Head of Department',
            approval_status='Approved'
        ).exists()

        leave_requests_data.append({
            'id': leave.id,
            'employee_name': leave.employee.name,
            'start_date': leave.start_date,
            'end_date': leave.end_date,
            'status': "Approved by Head" if head_approved else "Pending",
            'download_enabled': head_approved,
        })

    paginator = Paginator(leave_requests_data, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'sp/leave_requests_table.html', {'page_obj': page_obj})


def leave_requests_table_hd(request):
    leave_requests = LeaveRequest.objects.all()

    leave_requests_data = []
    for leave in leave_requests:
        head_approved = leave.leaveapproval_set.filter(
            approver_role__role_name='Head of Department',
            approval_status='Approved'
        ).exists()

        leave_requests_data.append({
            'id': leave.id,
            'employee_name': leave.employee.name,
            'start_date': leave.start_date,
            'end_date': leave.end_date,
            'status': "Approved by Head" if head_approved else "Pending",
            'download_enabled': head_approved,
        })

    paginator = Paginator(leave_requests_data, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'hd/leave_requests_table.html', {'page_obj': page_obj})





from django.shortcuts import render
from .models import LeaveRequest, LeaveApproval
from django.contrib.auth.decorators import login_required


from django.shortcuts import render
from .models import LeaveRequest, LeaveApproval



from django.shortcuts import render
from .models import LeaveRequest, LeaveApproval
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from .models import LeaveRequest
from django.core.paginator import Paginator
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .models import LeaveRequest

@login_required
def hr_dashboard(request):
    leave_requests = LeaveRequest.objects.all()  # Or filter as needed

    # Pre-fetching approvals related to the leave requests to optimize database hits
    for leave_request in leave_requests:
        leave_request.hr_approval = leave_request.leaveapproval_set.filter(approver_role__role_name='HR').first()
        leave_request.supervisor_approval = leave_request.leaveapproval_set.filter(approver_role__role_name='Supervisor').first()
        leave_request.head_approval = leave_request.leaveapproval_set.filter(approver_role__role_name='Head of Department').first()
    
    paginator = Paginator(leave_requests, 10)  # Show 10 leave requests per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'hr/hr_dashboard.html', {'leave_requests': leave_requests, "page_obj": page_obj})


@login_required
def supervisor_dashboard(request):
    leave_requests = LeaveRequest.objects.filter(status='HR Approved')  # Only show HR-approved requests
    paginator = Paginator(leave_requests, 10)  # Show 10 leave requests per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'sp/supervisor_dashboard.html', {'page_obj': page_obj})


@login_required
def head_of_section_dashboard(request):
    leave_requests = LeaveRequest.objects.filter(status='Supervisor Approved')  # Only show Supervisor-approved requests
    paginator = Paginator(leave_requests, 10)  # Show 10 leave requests per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'hd/head_of_section_dashboard.html', {'page_obj': page_obj})


from django.shortcuts import render, redirect
from django.contrib.auth import login, authenticate
from django.contrib import messages
from .forms import RegistrationForm
from .models import Employee, Role
from django.contrib.auth.models import User
def register(request):
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            # Check if the username, email, employment number, or bank account number already exists
            username = form.cleaned_data['username']
            email = form.cleaned_data['email']
            employment_no = form.cleaned_data['employment_no']
            bank_account_no = form.cleaned_data['bank_account_no']
            
            # Validation for existing fields
            if User.objects.filter(username=username).exists():
                form.add_error('username', 'Username already exists.')
            if User.objects.filter(email=email).exists():
                form.add_error('email', 'Email already exists.')
            if Employee.objects.filter(employment_no=employment_no).exists():
                form.add_error('employment_no', 'Employment number already exists.')
            if Employee.objects.filter(bank_account_no=bank_account_no).exists():
                form.add_error('bank_account_no', 'Bank account number already exists.')

            if form.errors:
                messages.error(request, 'There was an error in the form. Please check your input.')
            else:
                # Create the user and employee records if all fields are unique
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=form.cleaned_data['password']
                )

                role = Role.objects.get(role_name=form.cleaned_data['role'])
                employee = Employee.objects.create(
                    user=user,
                    name=form.cleaned_data['name'],
                    grade=form.cleaned_data['grade'],
                    post=role,
                    employment_no=form.cleaned_data['employment_no'],
                    contact_address=form.cleaned_data['contact_address'],
                    bank_name=form.cleaned_data['bank_name'],
                    bank_account_no=form.cleaned_data['bank_account_no'],
                    annual_leave_entitlement=form.cleaned_data['annual_leave_entitlement']
                )

                # Log the user in and redirect to the appropriate dashboard
                login(request, user)
                messages.success(request, 'Registration successful! You are now logged in.')
                return redirect('employee_list')
    else:
        form = RegistrationForm()

    return render(request, 'admin_z/register.html', {'form': form})
from django.http import JsonResponse
from .models import User, Employee

def check_field_availability(request):
    field = request.GET.get('field')
    value = request.GET.get('value')
    exists = False

    if field == 'username':
        exists = User.objects.filter(username=value).exists()
    elif field == 'email':
        exists = User.objects.filter(email=value).exists()
    elif field == 'employment_no':
        exists = Employee.objects.filter(employment_no=value).exists()
    elif field == 'bank_account_no':
        exists = Employee.objects.filter(bank_account_no=value).exists()

    return JsonResponse({'exists': exists})


# views.py
# from .forms import EmailAuthenticationForm

# def custom_login(request):
#     if request.method == 'POST':
#         form = EmailAuthenticationForm(request.POST)
#         if form.is_valid():
#             email = form.cleaned_data['email']
#             password = form.cleaned_data['password']
#             user = authenticate(request, username=email, password=password)

#             if user is not None:
#                 login(request, user)
#                 messages.success(request, 'You have successfully logged in!')
#                 # Redirect user based on role
#                 if user.employee.grade == 'A' or user.employee.grade == 'B':
#                     return redirect('hr_dashboard')
#                 elif user.employee.grade == 'C':
#                     return redirect('supervisor_dashboard')
#                 elif user.employee.grade == 'E (P4)':
#                     return redirect('head_of_section_dashboard')
#                 else:
#                     return redirect('home')  # Or any default page

#             else:
#                 messages.error(request, 'Invalid email or password.')
#         else:
#             messages.error(request, 'Invalid form submission.')

#     else:
#         form = EmailAuthenticationForm()

#     return render(request, 'login.html', {'form': form})

from django.shortcuts import render
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.shortcuts import render, redirect
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from .EmailBackEnd import EmailBackEnd
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth import login as auth_login


from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.shortcuts import redirect, render

# def login_view(request):
#     if request.method == "POST":
#         user = EmailBackEnd.authenticate(request, username=request.POST.get('email'), password=request.POST.get('password'))
#         if user is not None:
#             login(request, user)
#             messages.success(request, 'You have successfully logged in!')
            
#             # Redirect user based on role
#             if user.employee.post and user.employee.post.role_name == 'HR':
#                 return redirect('hr_dashboard')
#             elif user.employee.post and user.employee.post.role_name == 'Supervisor':
#                 return redirect('supervisor_dashboard')
#             elif user.employee.post and user.employee.post.role_name == 'Head of Department':
#                 return redirect('head_of_section_dashboard')

#             elif user.employee.post and user.employee.post.role_name == 'Employee':
#                 return redirect('homepage')

#             else:
#                 return redirect('admin_dash')  # Default page for employees with no specific role
                
#         else:
#             messages.error(request, "Invalid Login Credentials!")
#             return render(request, 'login.html')
#     else:
#         return render(request, 'login.html')

from django.shortcuts import redirect
from django.contrib.auth import login
from django.contrib import messages
from .models import Employee

def login_view(request):
    if request.method == "POST":
        user = EmailBackEnd.authenticate(request, username=request.POST.get('email'), password=request.POST.get('password'))
        if user is not None:
            login(request, user)
            messages.success(request, 'You have successfully logged in!')
            
            # Check if the user has an associated employee record
            try:
                employee = user.employee
                if employee.post and employee.post.role_name == 'HR':
                    return redirect('hr_dashboard')
                elif employee.post and employee.post.role_name == 'Supervisor':
                    return redirect('supervisor_dashboard')
                elif employee.post and employee.post.role_name == 'Head of Department':
                    return redirect('head_of_section_dashboard')
                elif employee.post and employee.post.role_name == 'Employee':
                    return redirect('homepage')
                elif employee.post and employee.post.role_name == 'System Administrator':
                    return redirect('ad_dash')

            except Employee.DoesNotExist:
                # If the user doesn't have an associated Employee record, redirect to the Django admin dashboard
                return redirect('/admin/')  # Redirect to Django admin interface

        else:
            messages.error(request, "Invalid Login Credentials!")
            return render(request, 'login.html')
    else:
        return render(request, 'login.html')


def logoutuser(request):
    logout(request)
    return redirect('login_view')

def success_page(request):
    return render(request, 'success.html')


from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa  # Use any library for PDF generation

# @login_required
# def generate_leave_form_view(request, leave_id):
#     leave_request = LeaveRequest.objects.get(id=leave_id)

#     # Fetch details for the leave form
#     context = {
#         'name': leave_request.employee.name,
#         'grade': leave_request.employee.grade,
#         'employment_no': leave_request.employee.employment_no,
#         'bank_name': leave_request.employee.bank_name,
#         'account_no': leave_request.employee.bank_account_no,
#         'start_date': leave_request.start_date,
#         'end_date': leave_request.end_date,
#         'contact_address': leave_request.employee.contact_address,  # Access through the employee object
#         'days_remaining': leave_request.days_remaining(),  # Use the method here
#         'hr_approval': leave_request.leaveapproval_set.filter(approver_role__role_name='HR').first(),
#         'supervisor_approval': leave_request.leaveapproval_set.filter(approver_role__role_name='Supervisor').first(),
#         'head_approval': leave_request.leaveapproval_set.filter(approver_role__role_name='Head of Section').first(),
#     }

#     # Render the template to HTML
#     template = get_template('leave_form_pdf.html')
#     html = template.render(context)

#     # Convert HTML to PDF
#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = f'attachment; filename="Leave_Form_{leave_id}.pdf"'
#     pisa_status = pisa.CreatePDF(html, dest=response)

#     if pisa_status.err:
#         return HttpResponse("Error generating PDF", status=500)

#     return response

from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from xhtml2pdf import pisa
from django.template.loader import get_template

import base64
from django.utils.html import escape

def generate_leave_form_view(request, leave_id):
    leave_request = get_object_or_404(LeaveRequest, id=leave_id)

    # Fetch approvals
    hr_approval = leave_request.leaveapproval_set.filter(approver_role__role_name='HR').first()
    supervisor_approval = leave_request.leaveapproval_set.filter(approver_role__role_name='Supervisor').first()
    head_approval = leave_request.leaveapproval_set.filter(approver_role__role_name='Head of Department').first()

    # Helper function to convert an image to Base64
    def image_to_base64(image_field):
        if image_field:
            try:
                with open(image_field.path, "rb") as img_file:
                    return f"data:image/jpeg;base64,{base64.b64encode(img_file.read()).decode()}"
            except Exception:
                return None
        return None

    # Convert signatures to Base64
    hr_signature_base64 = image_to_base64(hr_approval.signature) if hr_approval and hr_approval.signature else None
    supervisor_signature_base64 = image_to_base64(supervisor_approval.signature) if supervisor_approval and supervisor_approval.signature else None
    head_signature_base64 = image_to_base64(head_approval.signature) if head_approval and head_approval.signature else None

    # Prepare context for rendering the PDF
    context = {
        'name': leave_request.employee.name,
        'grade': leave_request.employee.get_grade_display(),
        'post': leave_request.employee.post.role_name if leave_request.employee.post else "Not Assigned",
        'employment_no': leave_request.employee.employment_no,
        'bank_name': leave_request.employee.bank_name,
        'account_no': leave_request.employee.bank_account_no,
        'start_date': leave_request.start_date,
        'end_date': leave_request.end_date,
        'contact_address': leave_request.contact_address_during_leave or "Not Provided",
        'days_remaining': leave_request.days_remaining(),
        'leave_grant': leave_request.leave_grant_requested or "Not Requested",
        'application_date': leave_request.submission_date.strftime("%Y-%m-%d"),
        'hr_signature': hr_signature_base64 or "Not Provided",
        'hr_date': hr_approval.action_date.strftime("%Y-%m-%d") if hr_approval else "Not Provided",
        'supervisor_signature': supervisor_signature_base64 or "Not Provided",
        'supervisor_date': supervisor_approval.action_date.strftime("%Y-%m-%d") if supervisor_approval else "Not Provided",
        'head_signature': head_signature_base64 or "Not Provided",
        'head_date': head_approval.action_date.strftime("%Y-%m-%d") if head_approval else "Not Provided",
    }

    # Render template into HTML
    template = get_template('leave_form_pdf.html')
    html = template.render(context)

    # Convert HTML to PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Leave_Form_{leave_id}.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse("Error generating PDF", status=500)

    return response




# from django.shortcuts import redirect
# from .models import LeaveRequest, LeaveApproval, Employee

# def leave_request_handler(request):
#     if request.method == 'POST':
#         employee_id = request.POST['employee']
#         employee = Employee.objects.get(id=employee_id)
        
#         # Create Leave Request
#         leave_request = LeaveRequest.objects.create(
#             employee=employee,
#             start_date=request.POST['start_date'],
#             end_date=request.POST['end_date'],
#             number_of_days=request.POST['number_of_days'],
#             contact_address_during_leave=request.POST.get('contact_address_during_leave', ''),
#             leave_grant_requested=request.POST.get('leave_grant_requested', '')
#         )
        
#         # Add Approval
#         approver_id = request.POST['approver']
#         approver = Employee.objects.get(id=approver_id)
#         approval_status = request.POST['approval_status']
#         comments = request.POST.get('comments', '')
        
#         LeaveApproval.objects.create(
#             leave_request=leave_request,
#             approver=approver,
#             approver_role=approver.post,
#             approval_status=approval_status,
#             comments=comments
#         )
        
#         return redirect('hr_dashboard')  # Redirect after successful submission
#     return redirect('leave_request_form')

from django.shortcuts import redirect
from .models import LeaveRequest, LeaveApproval, Employee

def leave_request_handler(request):
    if request.method == 'POST':
        employee_id = request.POST['employee']
        employee = Employee.objects.get(id=employee_id)
        
        # Create Leave Request
        leave_request = LeaveRequest.objects.create(
            employee=employee,
            start_date=request.POST['start_date'],
            end_date=request.POST['end_date'],
            number_of_days=request.POST['number_of_days'],
            contact_address_during_leave=request.POST.get('contact_address_during_leave', ''),
            leave_grant_requested=request.POST.get('leave_grant_requested', '')
        )
        
        # Add Approval
        approver_id = request.POST['approver']
        approver = Employee.objects.get(id=approver_id)
        approval_status = request.POST['approval_status']
        comments = request.POST.get('comments', '')
        
        # Handle the signature upload
        signature = request.FILES.get('signature')  # Get the uploaded signature image
        
        LeaveApproval.objects.create(
            leave_request=leave_request,
            approver=approver,
            approver_role=approver.post,
            approval_status=approval_status,
            comments=comments,
            signature=signature  # Save the signature if it was uploaded
        )
        
        return redirect('hr_dashboard')  # Redirect after successful submission
    
    return redirect('leave_request_form')


from django.shortcuts import render
from .models import Employee, Role

def leave_request_view(request):
    employees = Employee.objects.all()  # All employees
    approvers = Employee.objects.filter(post__role_name__in=['HR', 'Supervisor', 'Head of Department'])  # Approvers
    
    return render(request, 'hr/create.html', {
        'employees': employees,
        'approvers': approvers,
    })


from django.shortcuts import render
from .models import Employee

from django.shortcuts import render
from .models import Employee

def employee_list(request):
    employees = Employee.objects.select_related('user', 'post').all()
    for employee in employees:
        employee.role_name = employee.post.role_name if employee.post else 'N/A'

    paginator = Paginator(employees, 10)  # Show 10 leave requests per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        "page_obj" : page_obj

    }
    return render(request, 'admin_z/emp.html', context)


from django.shortcuts import get_object_or_404, redirect

def delete_employee(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)
    employee.user.delete()  # This will delete the user as well
    employee.delete()
    return redirect('employee_list')

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.models import User
from .models import Employee
from django.contrib.auth.hashers import make_password

def edit_employee(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)

    if request.method == 'POST':
        employee.name = request.POST['name']
        employee.grade = request.POST['grade']
        employee.employment_no = request.POST['employment_no']
        employee.contact_address = request.POST['contact_address']
        employee.bank_name = request.POST['bank_name']
        employee.bank_account_no = request.POST['bank_account_no']

        # Change Password
        if request.POST.get('password'):
            employee.user.password = make_password(request.POST['password'])
            employee.user.save()

        employee.save()
        return redirect('employee_list')

    return render(request, 'admin_z/edit_employee.html', {'employee': employee})


from django.shortcuts import render, redirect
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .forms import UserUpdateForm, PasswordChangeForm

@login_required
def user_profile(request):
    if request.method == 'POST':
        # Handling user info update (username, email)
        if 'update_user_info' in request.POST:
            user_form = UserUpdateForm(request.POST, instance=request.user)
            if user_form.is_valid():
                user_form.save()
                messages.success(request, 'Your account information has been updated.')
                return redirect('user_profile')
        # Handling password change
        elif 'change_password' in request.POST:
            password_form = PasswordChangeForm(request.POST)
            if password_form.is_valid():
                old_password = password_form.cleaned_data['old_password']
                new_password = password_form.cleaned_data['new_password']

                if not request.user.check_password(old_password):
                    password_form.add_error('old_password', 'Current password is incorrect.')
                else:
                    request.user.set_password(new_password)
                    request.user.save()
                    update_session_auth_hash(request, request.user)  # Prevent logging the user out
                    messages.success(request, 'Your password has been updated successfully.')
                    return redirect('user_profile')
            else:
                password_form.add_error(None, 'There was an error with your password change request.')

    else:
        user_form = UserUpdateForm(instance=request.user)
        password_form = PasswordChangeForm()

    return render(request, 'user_profile.html', {
        'user_form': user_form,
        'password_form': password_form


    })




def user_profile_hr(request):
    if request.method == 'POST':
        # Handling user info update (username, email)
        if 'update_user_info' in request.POST:
            user_form = UserUpdateForm(request.POST, instance=request.user)
            if user_form.is_valid():
                user_form.save()
                messages.success(request, 'Your account information has been updated.')
                return redirect('user_profile_hr')
        # Handling password change
        elif 'change_password' in request.POST:
            password_form = PasswordChangeForm(request.POST)
            if password_form.is_valid():
                old_password = password_form.cleaned_data['old_password']
                new_password = password_form.cleaned_data['new_password']

                if not request.user.check_password(old_password):
                    password_form.add_error('old_password', 'Current password is incorrect.')
                else:
                    request.user.set_password(new_password)
                    request.user.save()
                    update_session_auth_hash(request, request.user)  # Prevent logging the user out
                    messages.success(request, 'Your password has been updated successfully.')
                    return redirect('user_profile_hr')
            else:
                password_form.add_error(None, 'There was an error with your password change request.')

    else:
        user_form = UserUpdateForm(instance=request.user)
        password_form = PasswordChangeForm()

    return render(request, 'hr/user_profile.html', {
        'user_form': user_form,
        'password_form': password_form

        
    })


def user_profile_sp(request):
    if request.method == 'POST':
        # Handling user info update (username, email)
        if 'update_user_info' in request.POST:
            user_form = UserUpdateForm(request.POST, instance=request.user)
            if user_form.is_valid():
                user_form.save()
                messages.success(request, 'Your account information has been updated.')
                return redirect('user_profile_sp')
        # Handling password change
        elif 'change_password' in request.POST:
            password_form = PasswordChangeForm(request.POST)
            if password_form.is_valid():
                old_password = password_form.cleaned_data['old_password']
                new_password = password_form.cleaned_data['new_password']

                if not request.user.check_password(old_password):
                    password_form.add_error('old_password', 'Current password is incorrect.')
                else:
                    request.user.set_password(new_password)
                    request.user.save()
                    update_session_auth_hash(request, request.user)  # Prevent logging the user out
                    messages.success(request, 'Your password has been updated successfully.')
                    return redirect('user_profile_sp')
            else:
                password_form.add_error(None, 'There was an error with your password change request.')

    else:
        user_form = UserUpdateForm(instance=request.user)
        password_form = PasswordChangeForm()

    return render(request, 'sp/user_profile.html', {
        'user_form': user_form,
        'password_form': password_form

        
    })

def user_profile_hd(request):
    if request.method == 'POST':
        # Handling user info update (username, email)
        if 'update_user_info' in request.POST:
            user_form = UserUpdateForm(request.POST, instance=request.user)
            if user_form.is_valid():
                user_form.save()
                messages.success(request, 'Your account information has been updated.')
                return redirect('user_profile_hd')
        # Handling password change
        elif 'change_password' in request.POST:
            password_form = PasswordChangeForm(request.POST)
            if password_form.is_valid():
                old_password = password_form.cleaned_data['old_password']
                new_password = password_form.cleaned_data['new_password']

                if not request.user.check_password(old_password):
                    password_form.add_error('old_password', 'Current password is incorrect.')
                else:
                    request.user.set_password(new_password)
                    request.user.save()
                    update_session_auth_hash(request, request.user)  # Prevent logging the user out
                    messages.success(request, 'Your password has been updated successfully.')
                    return redirect('user_profile_hd')
            else:
                password_form.add_error(None, 'There was an error with your password change request.')

    else:
        user_form = UserUpdateForm(instance=request.user)
        password_form = PasswordChangeForm()

    return render(request, 'hd/user_profile.html', {
        'user_form': user_form,
        'password_form': password_form

        
    })


import openpyxl
from django.http import HttpResponse
from .models import Employee

def export_employees_to_excel(request):
    # Create a workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"

    # Define the header row
    headers = [
        "Name",
        "Grade",
        "Post",
        "Employment No",
        "Contact Address",
        "Bank Name",
        "Bank Account No",
        "Annual Leave Entitlement",
        "Leave Days Taken",
    ]
    ws.append(headers)

    # Query the Employee model and populate the worksheet
    for employee in Employee.objects.all():
        row = [
            employee.name,
            employee.get_grade_display(),
            employee.post.role_name if employee.post else "N/A",
            employee.employment_no,
            employee.contact_address,
            employee.bank_name,
            employee.bank_account_no,
            employee.annual_leave_entitlement,
            employee.leave_days_taken,
        ]
        ws.append(row)

    # Set up the HTTP response with content disposition
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="employees.xlsx"'

    # Save the workbook to the response
    wb.save(response)

    return response


import openpyxl
from django.http import HttpResponse
from .models import LeaveRequest, LeaveApproval

def export_leave_requests_to_excel(request):
    # Create a workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leave Requests"

    # Define the header row
    headers = [
        "Employee Name",
        "Start Date",
        "End Date",
        "Number of Days",
        "Contact Address",
        "Leave Grant Requested",
        "Status",
        "Submission Date",
        "Days Remaining",
    ]
    ws.append(headers)

    # Query the LeaveRequest model and populate the worksheet
    for leave_request in LeaveRequest.objects.all():
        row = [
            leave_request.employee.name,
            leave_request.start_date,
            leave_request.end_date,
            leave_request.number_of_days,
            leave_request.contact_address_during_leave or "N/A",
            leave_request.leave_grant_requested or "N/A",
            leave_request.status,
            leave_request.submission_date.strftime('%Y-%m-%d %H:%M:%S'),
            leave_request.days_remaining(),
        ]
        ws.append(row)

    # Set up the HTTP response with content disposition
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="leave_requests.xlsx"'

    # Save the workbook to the response
    wb.save(response)

    return response


def export_leave_approvals_to_excel(request):
    # Create a workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leave Approvals"

    # Define the header row
    headers = [
        "Leave Request ID",
        "Employee Name",
        "Approver Name",
        "Approver Role",
        "Approval Status",
        "Comments",
        "Action Date",
    ]
    ws.append(headers)

    # Query the LeaveApproval model and populate the worksheet
    for approval in LeaveApproval.objects.all():
        row = [
            approval.leave_request.id,
            approval.leave_request.employee.name,
            approval.approver.name,
            approval.approver_role.role_name,
            approval.approval_status,
            approval.comments or "N/A",
            approval.action_date.strftime('%Y-%m-%d %H:%M:%S'),
        ]
        ws.append(row)

    # Set up the HTTP response with content disposition
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="leave_approvals.xlsx"'

    # Save the workbook to the response
    wb.save(response)

    return response


import pandas as pd
from django.shortcuts import render, redirect
from django.contrib import messages
from .forms import ExcelUploadForm
from .models import Role, Employee

def upload_employees_excel(request):
    if request.method == "POST":
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES["file"]
            try:
                # Read the Excel file
                data = pd.read_excel(excel_file)

                # Validate columns
                required_columns = ["Name", "Grade", "Role", "Employment No", "Contact Address", "Bank Name", "Bank Account No", "Annual Leave Entitlement"]
                for column in required_columns:
                    if column not in data.columns:
                        raise ValueError(f"Missing required column: {column}")

                # Process rows
                for _, row in data.iterrows():
                    # Create or fetch Role
                    role, _ = Role.objects.get_or_create(role_name=row["Role"])

                    # Create Employee
                    Employee.objects.create(
                        name=row["Name"],
                        grade=row["Grade"],
                        post=role,
                        employment_no=row["Employment No"],
                        contact_address=row["Contact Address"],
                        bank_name=row["Bank Name"],
                        bank_account_no=row["Bank Account No"],
                        annual_leave_entitlement=int(row["Annual Leave Entitlement"]),
                        leave_days_taken=int(row.get("Leave Days Taken", 0)),
                    )

                messages.success(request, "Employees imported successfully!")
                return redirect("upload_employees_excel")
            except Exception as e:
                messages.error(request, f"Error processing the file: {e}")
    else:
        form = ExcelUploadForm()

    return render(request, "upload.html", {"form": form})
